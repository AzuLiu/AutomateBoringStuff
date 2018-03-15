import openpyxl
from openpyxl import Workbook
import os
# 1.确定详细品牌名
# 2.写入抬头信息
# 3.写入单价

# 1.Band
bandStr = '''大自然-佛山
东鹏-佛山洁具
奥普
贝朗
大卫
大自然-湖州
好门面
开来
科勒
美标
美赫
梦天
名门
欧派-木门
欧派-五金
普乐美
西顿-开关面板
西顿-浴霸
亚美利加
意利宝
优丽斯-南京
友邦
中迅
邦克
汇泰龙
尚高
大自然-上饶
优丽斯-重庆'''
bandList = bandStr.split('\n')
# bandTest = ['大自然-佛山', '东鹏-佛山洁具']
columnPrice = 4
columnSheet = 9
rate = 0.7
rowSale = 5

wbName = input('工作表路径：')
wb = openpyxl.load_workbook(wbName)
model = wb.get_sheet_by_name('Model')
price = wb.get_sheet_by_name('Price')
info = wb.get_sheet_by_name('Info')

# 2.Head and 3.Price
for band in bandList:
    columnPrice += 1
    sheet = wb.copy_worksheet(model)
    sheet.title = band + '重货抛货配置'
    sheet['A1'].value = band + '供应商价格策略配置-0001'
    sheet['B2'].value = band + '分站物流结算价'
    sheet['D2'].value = "=VLOOKUP(\"%s\",'Info'!$A$4:$M$52,2,0)" % band
    sheet['F2'].value = "=VLOOKUP(\"%s\",'Info'!$A$4:$M$52,11,0)"% band
    for rowNum in range(6, 848):
        rowNum2 = rowNum-2
        eachPrice = sheet.cell(row=rowNum, column=columnSheet).value = price.cell(row=rowNum2, column=columnPrice).value
        #sheet.cell(row=rowSale,column=columnSheet+1).value = '销售价/￥'
        #sheet.cell(row=rowNum, column=columnSheet+1).value = '%.2f'% (eachPrice/rate)

wb.save('F:\MyPython\priceTestX.xlsx')
